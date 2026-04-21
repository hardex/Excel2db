import re

import openpyxl
from typing import Any

from app.models.schemas import FieldModel
from app.services.logging_service import get_logger

_NEWLINE_RE = re.compile(r"[\r\n]+")


def _normalize_value(v):
    """Replace runs of CR/LF with a single space (mirrors browser single-line input paste)."""
    if isinstance(v, str):
        return _NEWLINE_RE.sub(" ", v)
    return v


def _apply_date_placeholder(value, value_type: str):
    """For date fields, treat the literal placeholder 'mm/yyyy' as empty (None)."""
    if value_type == "date" and isinstance(value, str) and value.strip().lower() == "mm/yyyy":
        return None
    return value

logger = get_logger()


def read_workbook_fields(file_path: str, fields: list[FieldModel]) -> dict[str, dict]:
    """Read values from an Excel workbook for the given fields.

    Each field's raw_cell_value flag controls whether to read raw (formulas as
    text) or calculated values.  The workbook is opened in both modes only when
    needed.
    """
    active_fields = [f for f in fields if f.active]
    needs_raw = any(f.raw_cell_value for f in active_fields)
    needs_labels = any(f.field_name_cell.strip() and f.field_name.strip() for f in active_fields)
    needs_calc = any(not f.raw_cell_value for f in active_fields) or needs_labels

    wb_raw = openpyxl.load_workbook(file_path, data_only=False) if needs_raw else None
    wb_calc = openpyxl.load_workbook(file_path, data_only=True) if needs_calc else None
    logger.info("Workbook opened successfully")

    results: dict[str, dict] = {}

    for field in active_fields:
        sheet_name = field.sheet
        cell_addr = field.cell
        wb = wb_raw if field.raw_cell_value else wb_calc

        if sheet_name not in wb.sheetnames:
            results[field.field_code] = {
                "value": None,
                "values": [],
                "error": f"Sheet '{sheet_name}' not found",
                "label_value": None,
            }
            logger.warning(
                f"Read field: field_code={field.field_code}, sheet={sheet_name}, "
                f"cell={cell_addr} — sheet not found"
            )
            continue

        ws = wb[sheet_name]
        try:
            values = _read_cell_spec(ws, cell_addr)
            values = [_apply_date_placeholder(v, field.value_type) for v in values]
            value = _collapse_values(values) if values else None
            result = {"value": value, "values": values, "error": None, "label_value": None}

            fnc = field.field_name_cell.strip()
            fn = field.field_name.strip()
            if fnc and fn and wb_calc is not None and sheet_name in wb_calc.sheetnames:
                try:
                    label_values = _read_cell_spec(wb_calc[sheet_name], fnc)
                    if label_values:
                        result["label_value"] = _collapse_values(label_values)
                except Exception as lexc:
                    logger.warning(
                        f"Label read failed: field_code={field.field_code}, "
                        f"field_name_cell={fnc} — {lexc}"
                    )

            results[field.field_code] = result
            logger.info(
                f"Read field: field_code={field.field_code}, sheet={sheet_name}, "
                f"cell={cell_addr}, raw={field.raw_cell_value}, n={len(values)}"
            )
        except Exception as exc:
            results[field.field_code] = {"value": None, "values": [], "error": str(exc), "label_value": None}
            logger.warning(
                f"Read field: field_code={field.field_code}, sheet={sheet_name}, "
                f"cell={cell_addr} — error: {exc}"
            )

    if wb_raw:
        wb_raw.close()
    if wb_calc:
        wb_calc.close()
    return results


def _read_cell_spec(ws, spec: str) -> list:
    """Read values matching a cell spec that may contain commas and ranges.

    Examples of accepted specs: "A1", "A1:B5", "A1, B5:C10, D2".
    """
    values: list = []
    for part in spec.split(","):
        p = part.strip()
        if not p:
            continue
        target = ws[p]
        if isinstance(target, tuple):  # range of rows
            for row in target:
                if isinstance(row, tuple):
                    for c in row:
                        values.append(_normalize_value(c.value))
                else:
                    values.append(_normalize_value(row.value))
        else:
            values.append(_normalize_value(target.value))
    return values


def _collapse_values(values: list):
    """Collapse a list of cell values to a single value for downstream consumers.

    Returns the single value if only one was read, otherwise a comma-joined string
    of stringified non-empty values.
    """
    if len(values) == 1:
        return values[0]
    return ", ".join("" if v is None else str(v) for v in values)


def read_single_cell(file_path: str, sheet_name: str, cell_addr: str, raw: bool = True, value_type: str = "") -> dict:
    """Read value(s) from a workbook (used by Test Cell feature).

    `cell_addr` may be a single cell (`A1`), a range (`A1:B5`), or a comma-
    separated list of cells/ranges (`A1, B5:C10`).

    Args:
        raw: If True, read raw values (formulas as text). If False, read calculated values.
        value_type: Optional field type. When "date", treats the literal "mm/yyyy"
            placeholder as empty (None).
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=not raw)
    except Exception as exc:
        return {"value": None, "values": [], "error": f"Cannot open workbook: {exc}"}

    if sheet_name not in wb.sheetnames:
        wb.close()
        return {"value": None, "values": [], "error": f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}"}

    ws = wb[sheet_name]
    try:
        values = _read_cell_spec(ws, cell_addr)
        values = [_apply_date_placeholder(v, value_type) for v in values]
        wb.close()
        if not values:
            return {"value": None, "values": [], "error": "No cells read"}
        return {"value": _collapse_values(values), "values": values, "error": None}
    except Exception as exc:
        wb.close()
        return {"value": None, "values": [], "error": str(exc)}


def get_sheet_names(file_path: str) -> list[str]:
    """Return list of sheet names from a workbook."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return []
