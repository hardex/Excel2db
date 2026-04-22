import json
import os
from pathlib import Path
from typing import Optional

import openpyxl

from app.models.schemas import TemplateModel
from app.services.model_service import ensure_model_has_fields

MAPPINGS_DIR = Path("mappings")
MAPPINGS_DIR.mkdir(parents=True, exist_ok=True)


def _template_filename(code: str, version: str) -> Path:
    return MAPPINGS_DIR / f"{code}_{version}.json"


def list_templates() -> list[TemplateModel]:
    """Return all templates sorted by code then version."""
    templates = []
    for path in sorted(MAPPINGS_DIR.glob("*.json")):
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            templates.append(TemplateModel(**data))
        except Exception:
            pass
    return templates


def get_template(code: str, version: str) -> Optional[TemplateModel]:
    path = _template_filename(code, version)
    if not path.exists():
        return None
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return TemplateModel(**data)


def get_default_template() -> Optional[TemplateModel]:
    templates = list_templates()
    for t in templates:
        if t.is_default:
            return t
    return templates[0] if templates else None


def save_template(template: TemplateModel) -> None:
    path = _template_filename(template.template_code, template.template_version)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(template.model_dump(), f, indent=2)
    # Auto-union this template's field_codes into the canonical model so
    # that later output always emits every field the model has ever seen.
    ensure_model_has_fields(
        template.template_code,
        [f.field_code for f in template.fields],
    )


def delete_template(code: str, version: str) -> bool:
    path = _template_filename(code, version)
    if path.exists():
        path.unlink()
        return True
    return False


def set_default_template(code: str, version: str) -> None:
    """Mark one template as default, clearing all others."""
    templates = list_templates()
    for t in templates:
        was_default = t.is_default
        t.is_default = t.template_code == code and t.template_version == version
        if was_default != t.is_default:
            save_template(t)
        elif t.template_code == code and t.template_version == version:
            save_template(t)


def template_exists(code: str, version: str) -> bool:
    return _template_filename(code, version).exists()


def detect_template_for_file(
    file_path: str,
) -> tuple[Optional[TemplateModel], dict]:
    """Auto-detect which template matches an uploaded Excel file.

    Each template may declare a `version_cell` (and optionally `version_cell_sheet`).
    We read that cell from the file and compare (trimmed string) to
    `model_version` — or to `template_version` when `model_version` is blank.

    Returns (template_or_None, info) where info describes the match attempt:
      - "candidates": [(code, version, sheet, cell, expected, actual)]
      - "error": optional string
    """
    info: dict = {"candidates": []}
    candidates = [t for t in list_templates() if t.version_cell.strip()]
    if not candidates:
        info["error"] = "No templates have a Version Cell configured."
        return None, info

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as exc:
        info["error"] = f"Cannot open workbook: {exc}"
        return None, info

    matches: list[TemplateModel] = []
    try:
        for tmpl in candidates:
            cell = tmpl.version_cell.strip()
            sheet_name = tmpl.version_cell_sheet.strip() or (wb.sheetnames[0] if wb.sheetnames else "")
            expected = (tmpl.model_version.strip() or tmpl.template_version.strip())
            actual = None
            try:
                if sheet_name in wb.sheetnames:
                    raw = wb[sheet_name][cell].value
                    actual = "" if raw is None else str(raw).strip()
            except Exception:
                actual = None
            info["candidates"].append({
                "code": tmpl.template_code,
                "version": tmpl.template_version,
                "sheet": sheet_name,
                "cell": cell,
                "expected": expected,
                "actual": actual,
            })
            if actual is not None and actual == expected and expected:
                matches.append(tmpl)
    finally:
        wb.close()

    if not matches:
        info["error"] = "No template matched the file's version cell."
        return None, info
    if len(matches) > 1:
        # Prefer the default among matches
        for m in matches:
            if m.is_default:
                return m, info
        # Otherwise first match
    return matches[0], info
